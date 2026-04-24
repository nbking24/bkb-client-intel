#!/usr/bin/env python3
"""
load-send-queue.py

Reads BKB-Send-Queue-Review.xlsx and upserts each sendable row into the
past_client_outreach table via the Client Hub bulk-load API.

Idempotent — safe to re-run after Nathan edits the spreadsheet. Existing
rows get updated with the latest text/notes but their stage and
timestamps are never reset.

Usage:
    python3 load-send-queue.py \\
        --xlsx "/path/to/BKB-Send-Queue-Review.xlsx" \\
        --api https://bkb-client-intel.vercel.app \\
        --token "$TICKET_AGENT_TOKEN"

Defaults:
    --xlsx: env PCO_XLSX_PATH, else ~/mnt/BKB/Marketing Project/BKB-Send-Queue-Review.xlsx
    --api:  env PCO_API_BASE,  else https://bkb-client-intel.vercel.app
    --token: env TICKET_AGENT_TOKEN (required if not in args)

Exit codes:
    0 = success
    1 = usage error
    2 = API error
    3 = file read error
"""
import argparse
import json
import os
import sys
import urllib.request
import urllib.error
from pathlib import Path

# Local config loader — adds CLI/env/~/.bkb-pco.env fallback
sys.path.insert(0, str(Path(__file__).parent))
import _config  # noqa: E402

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(3)

DEFAULT_XLSX = os.path.expanduser(
    "~/BKB/Marketing Project/BKB-Send-Queue-Review-LATEST.xlsx"
)

# LATEST file schema (19 columns) — see README in that workbook for details.
COLS = {
    "row": 1, "group": 2, "source": 3, "first_name": 4, "last_name": 5,
    "family": 6, "phone": 7, "need_loop": 8, "email": 9, "city": 10,
    "raw_project": 11, "descriptor": 12, "sendable": 13, "issue": 14,
    "char_count": 15, "final_text": 16, "commercial": 17,
    "origin_sheet": 18, "origin_row": 19,
}

# Group → priority (lower = higher priority)
#   5   = TEST — Nathan's known-safe contacts for pipeline validation; send first
#   10  = FRIEND / CUSTOM — new contacts (friends, subs, referral partners)
#   100 = INTRO / NO-INTRO — past clients
PRIORITY_BY_GROUP = {
    "TEST": 5,
    "FRIEND": 10,
    "CUSTOM": 10,
    "INTRO": 100,
    "NO-INTRO": 100,
}


def normalize_phone(phone):
    if not phone:
        return None
    s = str(phone).strip()
    # Strip Excel's scientific-notation .0 decimal
    if s.endswith(".0"):
        s = s[:-2]
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    return digits if len(digits) == 10 else None


def format_phone_display(phone_digits):
    if not phone_digits or len(phone_digits) != 10:
        return None
    return f"({phone_digits[:3]}) {phone_digits[3:6]}-{phone_digits[6:]}"


def map_source(val):
    """Map spreadsheet Source column to DB source values (check constraint = jt_past_project | loop_contact)."""
    if not val:
        return None
    v = str(val).lower()
    if "jt" in v or "jobtread" in v or "past project" in v:
        return "jt_past_project"
    if "new add" in v:
        # New Contact Adds — friends, subs, referral partners. DB only allows
        # the two original values, so bucket these under loop_contact.
        return "loop_contact"
    if "loop" in v or "ghl" in v:
        return "loop_contact"
    return None


def build_rows(ws):
    rows = []
    skipped = []
    for r in range(2, ws.max_row + 1):
        def get(key):
            col = COLS.get(key)
            if not col or col > ws.max_column:
                return None
            return ws.cell(r, col).value

        first_name = get("first_name")
        last_name = get("last_name")
        family = get("family")
        phone = get("phone")
        email = get("email")
        sendable = get("sendable")
        final_text = get("final_text")
        issue = get("issue")
        group = get("group")

        # Skip blank/separator rows
        if not (first_name or last_name or family) and not phone:
            continue

        # Determine inclusion stage
        explicit_skip = False
        if issue and "SKIP" in str(issue).upper():
            explicit_skip = True
        if str(sendable or "").upper() not in ("YES", ""):
            # Sendable=NO or any non-YES value → skip
            explicit_skip = True

        phone_digits = normalize_phone(phone)
        full_name = (family if family else f"{first_name or ''} {last_name or ''}".strip()) or None

        if not phone_digits and not explicit_skip:
            skipped.append({
                "row": r, "name": full_name, "reason": "no_phone"
            })
            continue

        priority = PRIORITY_BY_GROUP.get(str(group or "").upper(), 100)

        row_data = {
            "contact_key": phone_digits or f"row{r}-nophone",
            "first_name": (first_name or "").strip() if first_name else None,
            "last_name": (last_name or "").strip() if last_name else None,
            "full_name": full_name,
            "phone": format_phone_display(phone_digits) if phone_digits else None,
            "phone_digits": phone_digits,
            "email": email,
            "source": map_source(get("source")),
            "project_names": get("raw_project"),
            "city": get("city"),
            "priority": priority,
            "initial_text_body": final_text if sendable == "YES" and final_text else None,
            "flag_notes": issue or None,
        }

        if explicit_skip:
            row_data["stage"] = "skipped"

        rows.append(row_data)

    return rows, skipped


def post_bulk_load(api_base, token, rows):
    url = f"{api_base.rstrip('/')}/api/marketing/past-client/bulk-load"
    body = json.dumps({"rows": rows}).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=body,
        headers={
            "Content-Type": "application/json",
            "x-agent-token": token,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as res:
            return json.loads(res.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code}: {err_body}") from e


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx")
    ap.add_argument("--api")
    ap.add_argument("--token")
    ap.add_argument("--dry-run", action="store_true",
                    help="Parse the spreadsheet and print the payload without posting.")
    args = ap.parse_args()

    xlsx_path = _config.get_xlsx_path(cli_value=args.xlsx, default=DEFAULT_XLSX)
    api_base = _config.get_api_base(cli_value=args.api)
    token = args.token if args.dry_run else None
    if not args.dry_run:
        try:
            token = _config.get_token(cli_value=args.token)
        except RuntimeError as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)
    args.xlsx = xlsx_path
    args.api = api_base
    args.token = token

    path = Path(args.xlsx)
    if not path.exists():
        print(f"Error: xlsx not found at {path}", file=sys.stderr)
        sys.exit(3)

    wb = load_workbook(path)
    if "Send Queue" not in wb.sheetnames:
        print(f"Error: 'Send Queue' sheet not found. Sheets: {wb.sheetnames}",
              file=sys.stderr)
        sys.exit(3)

    rows, skipped = build_rows(wb["Send Queue"])
    print(f"Parsed {len(rows)} sendable rows, {len(skipped)} skipped for missing phone:")
    for s in skipped:
        print(f"  row {s['row']}: {s['name']} ({s['reason']})")

    if args.dry_run:
        print("\n--- DRY RUN — sample payload (first 2 rows) ---")
        print(json.dumps(rows[:2], indent=2))
        print(f"\n{len(rows)} rows would be posted to {args.api}")
        return

    # Post in chunks of 100 to stay comfortably under the 500 limit
    chunks = [rows[i:i + 100] for i in range(0, len(rows), 100)]
    totals = {"inserted": 0, "updated": 0, "skipped": 0, "errors": []}
    for i, chunk in enumerate(chunks, 1):
        print(f"Posting chunk {i}/{len(chunks)} ({len(chunk)} rows)…", end=" ")
        try:
            result = post_bulk_load(args.api, args.token, chunk)
        except Exception as e:
            print(f"\nAPI error on chunk {i}: {e}", file=sys.stderr)
            sys.exit(2)
        print(f"inserted={result.get('inserted', 0)} "
              f"updated={result.get('updated', 0)} "
              f"skipped={result.get('skipped', 0)} "
              f"errors={len(result.get('errors', []))}")
        totals["inserted"] += result.get("inserted", 0)
        totals["updated"] += result.get("updated", 0)
        totals["skipped"] += result.get("skipped", 0)
        totals["errors"].extend(result.get("errors", []))

    print(f"\n=== Done ===")
    print(f"Inserted: {totals['inserted']}")
    print(f"Updated:  {totals['updated']}")
    print(f"Skipped:  {totals['skipped']}")
    if totals["errors"]:
        print(f"Errors ({len(totals['errors'])}):")
        for e in totals["errors"][:10]:
            print(f"  {e}")


if __name__ == "__main__":
    main()
